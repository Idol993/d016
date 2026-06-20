from __future__ import annotations

import argparse
import json
import os
import sys
import logging
import yaml

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
from datetime import datetime
from typing import Optional

from models import (
    ApprovalRole,
    ApprovalStatus,
    DrillStatus,
    ReleaseRecord,
    ReleaseStatus,
    ReleaseType,
    HotfixMode,
    generate_id,
    now_iso,
    parse_iso,
)
from pre_check import PreCheckEngine
from approval import ApprovalEngine, ApprovalError
from gray_release import GrayReleaseEngine
from drill import DrillEngine
from report import ReportEngine
from audit import AuditLogger
from notify import NotifyEngine


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("main")


def load_config(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


class ReleasePlatform:
    def __init__(self, config_path: str = "config.yaml"):
        self.config = load_config(config_path)
        self.pre_check = PreCheckEngine(self.config)
        self.approval = ApprovalEngine(self.config)
        self.gray_release = GrayReleaseEngine(self.config)
        self.drill = DrillEngine(self.config)
        self.report = ReportEngine(self.config)
        self.audit = AuditLogger(self.config)
        self.notify = NotifyEngine(self.config)
        self.releases_dir = os.path.join(
            self.config.get("system", {}).get("data_dir", "./data"), "releases"
        )
        os.makedirs(self.releases_dir, exist_ok=True)

    def submit_release(
        self,
        version: str,
        previous_version: str,
        release_type: str,
        applicant: str,
        reason: str = "",
        hotfix_mode: Optional[str] = None,
    ) -> ReleaseRecord:
        rel_type = ReleaseType(release_type)
        release_id = generate_id("rel_")
        record = ReleaseRecord(
            release_id=release_id,
            version=version,
            previous_version=previous_version,
            release_type=rel_type,
            status=ReleaseStatus.PENDING_CHECK,
            applicant=applicant,
            apply_time=now_iso(),
            emergency_reason=reason if rel_type == ReleaseType.HOTFIX else "",
        )

        self.audit.log("submit_release", applicant, release_id, f"版本={version}, 类型={release_type}")
        self._save_release(record)
        logger.info("=" * 70)
        logger.info("发布申请已提交: %s, 版本: %s, 类型: %s", release_id, version, release_type)

        logger.info("-" * 70)
        logger.info("【阶段1】发布前置校验")
        pre_check_result = self.pre_check.run(release_id)
        record.pre_check_result = pre_check_result
        self._save_release(record)

        if not pre_check_result.passed:
            record.status = ReleaseStatus.CHECK_FAILED
            self._save_release(record)
            failed_items = [
                {"name": i.name, "message": i.message, "suggestion": i.suggestion, "value": i.value, "threshold": i.threshold}
                for i in pre_check_result.get_failed_items()
            ]
            self.notify.send_pre_check_failed(release_id, failed_items)
            self.audit.log("pre_check_failed", "system", release_id, f"未通过项: {len(failed_items)}")
            logger.error("前置校验未通过，发布已阻断，阻断原因如下:")
            self._print_pre_check_failed(failed_items)
            self._print_pre_check_result(pre_check_result)
            return record

        record.status = ReleaseStatus.CHECK_PASSED
        self._save_release(record)
        self.audit.log("pre_check_passed", "system", release_id)
        logger.info("前置校验全部通过 ✓")
        self._print_pre_check_result(pre_check_result)

        logger.info("-" * 70)
        logger.info("【阶段2】分级审批流转")
        record.status = ReleaseStatus.PENDING_APPROVAL
        flow = self.approval.create_flow(release_id, rel_type, reason, hotfix_mode)
        record.approval_flow = flow
        record.approval_first_at = now_iso()
        self._save_release(record)

        if flow.can_enter_gray() and rel_type == ReleaseType.HOTFIX and flow.hotfix_mode == HotfixMode.POST_SIGN:
            record.status = ReleaseStatus.APPROVAL_PASSED
            self._save_release(record)
            logger.info("紧急热修复(事后补签模式)，直接进入灰度发布")
        else:
            pending_roles = self.approval.get_next_pending_roles(release_id)
            if pending_roles:
                self.notify.send_approval_notification(
                    release_id,
                    [r.value for r in pending_roles],
                    rel_type.value,
                )

        self._print_approval_status(release_id)
        return record

    def approve_release(
        self,
        release_id: str,
        role: str,
        approver: str,
        comment: str = "",
    ) -> ReleaseRecord:
        record = self._load_release(release_id)
        if record is None:
            raise ValueError(f"发布记录不存在: {release_id}")

        try:
            flow = self.approval.approve(release_id, ApprovalRole(role), approver, comment)
        except ApprovalError as e:
            logger.error("审批失败: %s", str(e))
            print(f"\n❌ 审批失败: {e}")
            if hasattr(e, 'current_role') and e.current_role:
                role_map = {
                    "customs": "关务审批",
                    "operations": "运营审批",
                    "finance": "财务审批",
                    "tech": "技术审批",
                }
                print(f"👉 当前等待: {role_map.get(e.current_role, e.current_role)}({e.current_role})")
            raise

        record.approval_flow = flow
        record.approval_last_at = now_iso()
        if record.approval_first_at:
            first = parse_iso(record.approval_first_at)
            last = parse_iso(record.approval_last_at)
            record.approval_duration_minutes = round((last - first).total_seconds() / 60, 2)

        self.audit.log("approve", approver, release_id, f"角色={role}, 意见={comment}")
        self._save_release(record)

        if flow.can_enter_gray():
            record.status = ReleaseStatus.APPROVAL_PASSED
            self._save_release(record)
            if flow.is_fully_approved():
                self.audit.log("approval_completed", "system", release_id)
                logger.info("[%s] 全部审批通过 ✓", release_id)
            else:
                self.audit.log("approval_partial", "system", release_id, "关键角色已通过，进入灰度")
                post_sign_roles = flow.get_roles_needing_post_sign()
                logger.info(
                    "[%s] 关键角色审批通过，可进入灰度；待补签角色: %s",
                    release_id, [r.value for r in post_sign_roles],
                )
        else:
            pending = self.approval.get_next_pending_roles(release_id)
            self.notify.send_approval_notification(
                release_id,
                [r.value for r in pending],
                record.release_type.value,
            )

        self._print_approval_status(release_id)
        return record

    def post_sign(
        self,
        release_id: str,
        role: str,
        approver: str,
        comment: str = "",
    ) -> ReleaseRecord:
        record = self._load_release(release_id)
        if record is None:
            raise ValueError(f"发布记录不存在: {release_id}")

        try:
            flow = self.approval.post_sign(release_id, ApprovalRole(role), approver, comment)
        except ApprovalError as e:
            logger.error("补签失败: %s", str(e))
            print(f"\n❌ 补签失败: {e}")
            raise

        record.approval_flow = flow
        self.audit.log("post_sign", approver, release_id, f"角色={role}, 意见={comment}")
        self._save_release(record)

        if flow.is_fully_approved():
            self.audit.log("post_sign_completed", "system", release_id, "全部补签完成")
            logger.info("[%s] 全部补签完成 ✓", release_id)

        self._print_approval_status(release_id)
        return record

    def reject_release(
        self,
        release_id: str,
        role: str,
        approver: str,
        comment: str = "",
    ) -> ReleaseRecord:
        record = self._load_release(release_id)
        if record is None:
            raise ValueError(f"发布记录不存在: {release_id}")

        try:
            flow = self.approval.reject(release_id, ApprovalRole(role), approver, comment)
        except ApprovalError as e:
            logger.error("驳回失败: %s", str(e))
            print(f"\n❌ 驳回失败: {e}")
            raise

        record.approval_flow = flow
        record.status = ReleaseStatus.APPROVAL_REJECTED
        self.audit.log("reject", approver, release_id, f"角色={role}, 原因={comment}")
        self._save_release(record)
        logger.info("[%s] 审批已驳回", release_id)
        self._print_approval_status(release_id)
        return record

    def execute_gray_release(
        self,
        release_id: str,
        simulate_failure_tier: int = 0,
    ) -> dict:
        record = self._load_release(release_id)
        if record is None:
            raise ValueError(f"发布记录不存在: {release_id}")

        if record.status not in (ReleaseStatus.APPROVAL_PASSED, ReleaseStatus.GRAY_DEPLOYING):
            flow = self.approval.load_flow(release_id)
            reason = ""
            if flow and not flow.can_enter_gray():
                critical = [r.role.value for r in flow.records if r.is_critical and r.status.value == "pending"]
                if critical:
                    reason = f"关键角色尚未审批: {', '.join(critical)}"
            raise ValueError(f"发布状态不允许灰度: {record.status.value}. {reason}")

        record.status = ReleaseStatus.GRAY_DEPLOYING
        self._save_release(record)
        self.audit.log("gray_release_start", "system", release_id, f"版本={record.version}")

        logger.info("-" * 70)
        logger.info("【阶段3】口岸灰度发布与监控")
        result = self.gray_release.run_full_gray_release(
            release_id=release_id,
            version=record.version,
            previous_version=record.previous_version,
            simulate_failure_tier=simulate_failure_tier if simulate_failure_tier > 0 else None,
        )

        port_records = result.get("port_records", [])
        for pr in port_records:
            record.port_records.append(pr)
            record.port_results[pr["port_name"]] = pr["status"] != "rolled_back"

        if result.get("circuit_break"):
            record.status = ReleaseStatus.ROLLED_BACK
            cb_events = result.get("circuit_break_events", [])
            for cb in cb_events:
                record.circuit_break_events.append(cb)
            self._save_release(record)

            cb_data = result.get("results", [])[-1].get("monitor", {}).get("circuit_break_event", {})
            affected_ports = cb_data.get("affected_ports", [])
            self.notify.send_circuit_break_alert(
                release_id,
                cb_data.get("reason", ""),
                cb_data.get("trigger_value", 0),
                cb_data.get("threshold", 0),
                affected_ports,
                record.previous_version,
            )
            self.audit.log(
                "circuit_break_rollback",
                "system",
                release_id,
                f"熔断原因={cb_data.get('reason')}, 影响口岸={affected_ports}",
            )
            logger.warning("[%s] 熔断触发，已自动回滚至 %s", release_id, record.previous_version)
        else:
            record.status = ReleaseStatus.FULL_RELEASED
            record.finish_time = now_iso()
            self._save_release(record)
            self.audit.log("release_completed", "system", release_id, f"版本={record.version}")
            logger.info("[%s] 灰度发布全量完成，版本 %s 已上线 ✓", release_id, record.version)

        self._save_release(record)
        self._print_gray_result(result)
        return result

    def auto_approve_and_release(
        self,
        release_id: str,
    ) -> dict:
        record = self._load_release(release_id)
        if record is None:
            raise ValueError(f"发布记录不存在: {release_id}")
        if record.status not in (ReleaseStatus.CHECK_PASSED, ReleaseStatus.PENDING_APPROVAL):
            return {"success": False, "reason": f"状态不正确: {record.status.value}，无法自动审批"}

        flow = self.approval.load_flow(release_id)
        if flow is None:
            return {"success": False, "reason": "审批流不存在"}

        if flow.release_type == ReleaseType.REGULAR:
            role_name_map = {
                "customs": "关务审批",
                "operations": "运营审批",
                "finance": "财务审批",
                "tech": "技术审批",
            }
            logger.info("自动完成常规审批 (顺序: 关务 → 运营 → 财务 → 技术)")
            for role_enum in flow.records:
                if role_enum.status.value != "pending":
                    continue
                role = role_enum.role.value
                role_name = role_name_map.get(role, role)
                approver = f"{role}_auto"
                logger.info("  ▶ %s 审批中...", role_name)
                try:
                    record = self.approve_release(
                        release_id=release_id,
                        role=role,
                        approver=approver,
                        comment=f"自动审批 - {role_name}评估通过",
                    )
                except Exception as e:
                    return {"success": False, "reason": f"自动审批失败({role_name}): {e}"}
                if record.status == ReleaseStatus.APPROVAL_REJECTED:
                    return {"success": False, "reason": f"{role_name}审批被驳回"}
                logger.info("  ✅ %s 审批通过", role_name)

        record = self._load_release(release_id)
        if record is None or record.status != ReleaseStatus.APPROVAL_PASSED:
            current_status = record.status.value if record else "unknown"
            return {"success": False, "reason": f"审批未完成，当前状态: {current_status}"}

        logger.info("审批全部完成，开始灰度发布...")
        result = self.execute_gray_release(release_id)
        return {"success": True, "release_result": result}

    def run_drill(self, target_version: str = "v2.0.0", rollback_version: str = "v1.9.0") -> dict:
        logger.info("-" * 70)
        logger.info("【阶段4】回滚演练")
        drill_record = self.drill.schedule_drill(target_version, rollback_version)
        self.audit.log("drill_scheduled", "system", drill_record.drill_id, f"目标={target_version}")

        drill_record = self.drill.execute_drill(drill_record.drill_id)
        self.audit.log(
            "drill_completed",
            "system",
            drill_record.drill_id,
            f"状态={drill_record.status.value}, 耗时={drill_record.duration_seconds}s",
        )
        self.notify.send_drill_result(
            drill_record.drill_id,
            drill_record.status.value,
            drill_record.duration_seconds,
            drill_record.result_detail,
        )
        return {
            "drill_id": drill_record.drill_id,
            "status": drill_record.status.value,
            "duration_seconds": drill_record.duration_seconds,
            "detail": drill_record.result_detail,
        }

    def generate_report(self, week_offset: int = 0) -> dict:
        logger.info("-" * 70)
        logger.info("生成运营周报")
        report = self.report.generate_weekly_report(week_offset)
        self.audit.log("report_generated", "system", report.report_id)

        success_rate = (
            round(report.success_releases / report.total_releases, 4)
            if report.total_releases > 0 else 0.0
        )

        return {
            "report_id": report.report_id,
            "period": f"{report.period_start} ~ {report.period_end}",
            "total_releases": report.total_releases,
            "success_releases": report.success_releases,
            "rollback_count": report.rollback_count,
            "success_rate": success_rate,
            "avg_approval_duration_minutes": report.avg_approval_duration_minutes,
            "failure_rate_by_port": report.failure_rate_by_port,
            "file_paths": report.file_paths,
        }

    def query_release_history(
        self,
        version: str = "",
        port: str = "",
        start_date: str = "",
        end_date: str = "",
        release_id: str = "",
    ) -> list[dict]:
        return self.audit.query_release_history(
            version=version or None,
            port=port or None,
            start_date=start_date or None,
            end_date=end_date or None,
            release_id=release_id or None,
        )

    def export_release_history(
        self,
        output_path: str,
        version: str = "",
        port: str = "",
        start_date: str = "",
        end_date: str = "",
        release_id: str = "",
    ) -> str:
        return self.audit.export_release_history(
            output_path=output_path,
            version=version or None,
            port=port or None,
            start_date=start_date or None,
            end_date=end_date or None,
            release_id=release_id or None,
        )

    def query_audit(
        self,
        start_date: str = "",
        end_date: str = "",
        action: str = "",
        operator: str = "",
        target: str = "",
    ) -> list[dict]:
        return self.audit.query(
            start_date=start_date or None,
            end_date=end_date or None,
            action=action or None,
            operator=operator or None,
            target=target or None,
        )

    def verify_audit_integrity(self, date_str: str = "") -> dict:
        return self.audit.verify_integrity(date_str or None)

    def get_approval_status(self, release_id: str) -> dict:
        return self.approval.get_approval_status_summary(release_id)

    def get_release_summary(
        self,
        version: str = "",
        start_date: str = "",
        end_date: str = "",
    ) -> dict:
        """按版本/时间段汇总发布统计"""
        records = self.audit.query_release_history(
            version=version or None,
            start_date=start_date or None,
            end_date=end_date or None,
        )
        summary = {
            "total": len(records),
            "regular_count": 0,
            "hotfix_count": 0,
            "success_count": 0,
            "circuit_break_count": 0,
            "rollback_count": 0,
            "post_sign_total": 0,
            "post_sign_done": 0,
            "post_sign_pending": 0,
            "post_sign_rate": 0.0,
            "success_rate": 0.0,
            "avg_approval_minutes": 0.0,
            "health_score": 0,
            "health_label": "",
            "records_processed": len(records),
            "by_version": {},
            "start_date": start_date,
            "end_date": end_date,
            "filter_version": version,
        }
        approval_durations = []

        for rec in records:
            rtype = rec.get("release_type", "regular")
            if rtype == "regular":
                summary["regular_count"] += 1
            elif rtype == "hotfix":
                summary["hotfix_count"] += 1

            status = rec.get("status", "")
            if status in ("full_released", "approval_passed"):
                summary["success_count"] += 1
            elif status in ("circuit_break", "rolled_back"):
                summary["rollback_count"] += 1

            for cb in rec.get("circuit_break_events", []):
                if cb:
                    summary["circuit_break_count"] += 1

            approval = rec.get("approval", {})
            for ar in approval.get("records", []):
                is_critical = ar.get("is_critical", False)
                st = ar.get("status", "")
                if rtype == "hotfix" and not is_critical:
                    summary["post_sign_total"] += 1
                    if st == "post_signed":
                        summary["post_sign_done"] += 1
                    elif st in ("skipped", "pending"):
                        summary["post_sign_pending"] += 1

            dur = rec.get("approval_duration_minutes")
            if dur and dur > 0:
                approval_durations.append(dur)

            v = rec.get("version", "unknown")
            if v not in summary["by_version"]:
                summary["by_version"][v] = {
                    "count": 0, "success": 0, "rollback": 0,
                    "regular": 0, "hotfix": 0,
                }
            summary["by_version"][v]["count"] += 1
            if status in ("full_released", "approval_passed"):
                summary["by_version"][v]["success"] += 1
            elif status in ("circuit_break", "rolled_back"):
                summary["by_version"][v]["rollback"] += 1
            if rtype == "regular":
                summary["by_version"][v]["regular"] += 1
            elif rtype == "hotfix":
                summary["by_version"][v]["hotfix"] += 1

        if summary["total"] > 0:
            summary["success_rate"] = round(summary["success_count"] / summary["total"] * 100, 2)
        if summary["post_sign_total"] > 0:
            summary["post_sign_rate"] = round(summary["post_sign_done"] / summary["post_sign_total"] * 100, 2)
        if approval_durations:
            summary["avg_approval_minutes"] = round(sum(approval_durations) / len(approval_durations), 2)

        score = 100
        if summary["total"] > 0:
            score -= max(0, (100 - summary["success_rate"])) * 0.5
        score -= summary["circuit_break_count"] * 10
        score -= summary["rollback_count"] * 15
        if summary["post_sign_total"] > 0:
            score -= max(0, (100 - summary["post_sign_rate"])) * 0.3
        score = max(0, min(100, int(score)))
        summary["health_score"] = score

        if score >= 90:
            summary["health_label"] = "✅ 健康"
        elif score >= 70:
            summary["health_label"] = "⚠️  需关注"
        elif score >= 50:
            summary["health_label"] = "🔥 异常"
        else:
            summary["health_label"] = "💥 严重"

        return summary

    def get_port_history(self, port_name: str, limit: int = 10) -> dict:
        """按口岸聚合查看最近发布历史"""
        all_records = self.audit.query_release_history()
        port_records = []
        current_version = None
        current_time = ""

        for rec in all_records:
            for pr in rec.get("port_records", []):
                if not port_name or pr.get("port_name") == port_name:
                    cb = pr.get("circuit_break")
                    rollback_time = ""
                    if cb and cb.get("rollback_completed_at"):
                        rollback_time = cb["rollback_completed_at"]
                    elif pr.get("rollback_completed_at"):
                        rollback_time = pr["rollback_completed_at"]

                    release_time = pr.get("deploy_started_at", rec.get("apply_time", ""))
                    if release_time and release_time > current_time:
                        st = pr.get("status", "")
                        if st in ("released", "deployed", "monitoring"):
                            current_version = rec.get("version")
                            current_time = release_time
                        elif st in ("rolled_back", "circuit_break") and not current_version:
                            current_version = rec.get("previous_version")
                            current_time = rollback_time or release_time

                    port_records.append({
                        "release_id": rec.get("release_id"),
                        "version": rec.get("version"),
                        "previous_version": rec.get("previous_version"),
                        "release_type": rec.get("release_type"),
                        "tier": pr.get("tier"),
                        "status": pr.get("status"),
                        "release_time": release_time,
                        "circuit_break": cb is not None,
                        "circuit_break_reason": (cb or {}).get("reason", "") if cb else "",
                        "rollback_time": rollback_time,
                        "deploy_duration_minutes": 0,
                    })

        port_records.sort(key=lambda x: x["release_time"], reverse=True)
        port_records = port_records[:limit]

        has_risk = any(
            pr["status"] in ("circuit_break", "rolled_back")
            for pr in port_records[:3]
        ) if len(port_records) >= 3 else False

        return {
            "port_name": port_name,
            "total_records": len(port_records),
            "current_online_version": current_version,
            "current_deploy_time": current_time,
            "has_recent_risk": has_risk,
            "risk_level": "🔥 高风险" if has_risk else "✅ 正常",
            "records": port_records,
        }

    def check_compliance(self) -> dict:
        """合规校验：审批闭环、补签超期、顺序异常、回滚未复盘"""
        issues = []
        all_records = self.audit.query_release_history()
        now = parse_iso(now_iso())

        for rec in all_records:
            rid = rec.get("release_id")
            version = rec.get("version")
            rtype = rec.get("release_type")
            approval = rec.get("approval", {})
            records = approval.get("records", [])
            status = rec.get("status", "")

            if not records:
                continue

            approval_statuses = [r.get("status", "") for r in records]
            closed = all(s in ("approved", "post_signed", "rejected") for s in approval_statuses)

            if status in ("full_released", "gray_deploying", "check_passed", "pending_approval"):
                if not closed:
                    pending_roles = [
                        r["role"] for r in records
                        if r["status"] in ("pending", "skipped")
                    ]
                    issues.append({
                        "release_id": rid,
                        "version": version,
                        "issue_type": "审批未闭环",
                        "severity": "medium",
                        "detail": f"发布已进入{status}状态，但审批链路未完成",
                        "action": f"请处理待审批角色: {', '.join(pending_roles)}",
                    })

            if rtype == "hotfix":
                deadline_str = approval.get("post_sign_deadline", "")
                if deadline_str:
                    try:
                        deadline = parse_iso(deadline_str)
                        for r in records:
                            if not r.get("is_critical") and r.get("status") == "skipped":
                                if now > deadline:
                                    hours_overdue = round((now - deadline).total_seconds() / 3600, 1)
                                    issues.append({
                                        "release_id": rid,
                                        "version": version,
                                        "issue_type": "补签超期",
                                        "severity": "high",
                                        "detail": f"{r['role']}补签已超期 {hours_overdue} 小时",
                                        "action": f"立即完成 {r['role']} 事后补签",
                                    })
                    except Exception:
                        pass

            if rtype == "regular":
                last_approved_idx = -1
                for idx, r in enumerate(records):
                    s = r.get("status", "")
                    if s in ("approved", "post_signed"):
                        if idx <= last_approved_idx:
                            issues.append({
                                "release_id": rid,
                                "version": version,
                                "issue_type": "审批顺序异常",
                                "severity": "high",
                                "detail": f"{r['role']}在之前的角色未完成时已审批通过",
                                "action": "核查审批流程，确认是否存在越权审批",
                            })
                        last_approved_idx = idx

            if status in ("circuit_break", "rolled_back"):
                drill_count = 0
                try:
                    drill_records = self.drill.list_drills()
                    for d in drill_records:
                        if d.target_version == version and d.status in (DrillStatus.SUCCESS, DrillStatus.PASSED):
                            drill_count += 1
                except Exception:
                    pass
                if drill_count == 0:
                    issues.append({
                        "release_id": rid,
                        "version": version,
                        "issue_type": "回滚未复盘",
                        "severity": "medium",
                        "detail": "发生熔断/回滚后未进行演练复盘",
                        "action": "立即组织回滚演练并记录复盘结果",
                    })

        issues.sort(key=lambda x: {"high": 0, "medium": 1, "low": 2}[x.get("severity", "low")])
        return {
            "total_checked": len(all_records),
            "total_issues": len(issues),
            "high_count": sum(1 for i in issues if i["severity"] == "high"),
            "medium_count": sum(1 for i in issues if i["severity"] == "medium"),
            "low_count": sum(1 for i in issues if i["severity"] == "low"),
            "compliance_rate": round(
                max(0, (len(all_records) - sum(1 for i in issues if i["severity"] in ("high", "medium"))))
                / max(1, len(all_records)) * 100, 2
            ),
            "issues": issues,
        }

    def _save_release(self, record: ReleaseRecord):
        data = {
            "release_id": record.release_id,
            "version": record.version,
            "previous_version": record.previous_version,
            "release_type": record.release_type.value,
            "status": record.status.value,
            "applicant": record.applicant,
            "apply_time": record.apply_time,
            "finish_time": record.finish_time,
            "emergency_reason": record.emergency_reason,
            "approval_first_at": record.approval_first_at,
            "approval_last_at": record.approval_last_at,
            "approval_duration_minutes": record.approval_duration_minutes,
            "port_results": record.port_results,
        }
        path = os.path.join(self.releases_dir, f"{record.release_id}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _load_release(self, release_id: str):
        path = os.path.join(self.releases_dir, f"{release_id}.json")
        if not os.path.exists(path):
            return None
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        rec = ReleaseRecord(
            release_id=data["release_id"],
            version=data["version"],
            previous_version=data["previous_version"],
            release_type=ReleaseType(data["release_type"]),
            status=ReleaseStatus(data["status"]),
            applicant=data["applicant"],
            apply_time=data["apply_time"],
            finish_time=data.get("finish_time", ""),
            emergency_reason=data.get("emergency_reason", ""),
            approval_first_at=data.get("approval_first_at", ""),
            approval_last_at=data.get("approval_last_at", ""),
            approval_duration_minutes=data.get("approval_duration_minutes", 0.0),
        )
        rec.port_results = data.get("port_results", {})
        return rec

    def _print_pre_check_failed(self, failed_items: list[dict]):
        print()
        print("┌" + "─" * 68 + "┐")
        print("│" + " " * 20 + "⚠️  发布前置校验 - 阻断指标" + " " * 18 + "│")
        print("├──────────────────┬──────────┬──────────┬────────────────────────────────┤")
        print("│ 指标名称         │ 实际值   │ 阈值     │ 原因与建议                    │")
        print("├──────────────────┼──────────┼──────────┼────────────────────────────────┤")
        for item in failed_items:
            val = f"{item['value']:.2%}" if item["value"] is not None else "N/A"
            thr = f"{item['threshold']:.2%}" if item["threshold"] is not None else "N/A"
            print(f"│ {item['name']:<16} │ {val:<8} │ {thr:<8} │ {item['message'][:30]:<30} │")
            if item.get('suggestion'):
                print(f"│                  │          │          │   ▶ {item['suggestion'][:28]:<28} │")
        print("└──────────────────┴──────────┴──────────┴────────────────────────────────┘")
        print(f"\n❌ 共 {len(failed_items)} 项指标未达标，发布已阻断。请修复后重新提交。")
        print()

    def _print_pre_check_result(self, result):
        print()
        print("┌─────────────────────────────────────────────────────────────────────────┐")
        print("│                       发布前置校验结果                                  │")
        print("├──────────┬────────────────────┬────────┬────────────────────────────────┤")
        print("│ 分类     │ 检查项             │ 状态   │ 数值/说明                    │")
        print("├──────────┼────────────────────┼────────┼────────────────────────────────┤")
        for item in result.check_items:
            status_icon = "✓ PASS" if item.status.value == "pass" else "✗ FAIL"
            val_str = f"{item.value:.2%}" if item.value is not None else ""
            thr_str = f"(阈值 {item.threshold:.2%})" if item.threshold is not None else ""
            print(f"│ {item.category:<8} │ {item.name:<18} │ {status_icon:<6} │ {val_str + thr_str:<30} │")
        print("├──────────┴────────────────────┴────────┴────────────────────────────────┤")
        overall = "✓ 全部通过" if result.passed else "✗ 存在未通过项"
        print(f"│ 结论: {overall:<64} │")
        print("└─────────────────────────────────────────────────────────────────────────┘")

    def _print_approval_status(self, release_id: str):
        status = self.approval.get_approval_status_summary(release_id)
        if not status:
            return

        type_map = {"regular": "常规发布", "hotfix": "紧急热修复"}
        mode_map = {"parallel": "并行审批", "post_sign": "事后补签"}

        print()
        print("┌─────────────────────────────────────────────────────────────────────────┐")
        print("│                       审批流转状态                                      │")
        print("├─────────────────────────────────────────────────────────────────────────┤")
        print(f"│ 发布类型: {type_map.get(status['release_type'], status['release_type'])}", end="")
        if status.get('hotfix_mode'):
            mode = mode_map.get(status['hotfix_mode'], status['hotfix_mode'])
            print(f" (模式: {mode})")
        else:
            print()
        if status.get('emergency_reason'):
            print(f"│ 紧急原因: {status['emergency_reason'][:60]}")
        print(f"│ 关键角色已通过: {'是' if status['can_enter_gray'] else '否'}", end="")
        print(f"   全部审批通过: {'是 ✓' if status['is_fully_approved'] else '否'}")
        if status.get('post_sign_deadline'):
            print(f"│ 补签截止时间: {status['post_sign_deadline']}")
        print("├────────┬────────────────┬─────────┬─────────┬──────────────────────────┤")
        print("│ 序号   │ 角色           │ 关键    │ 状态    │ 审批人/意见               │")
        print("├────────┼────────────────┼─────────┼─────────┼──────────────────────────┤")

        role_names = {
            "customs": "关务审批", "operations": "运营审批",
            "finance": "财务审批", "tech": "技术审批",
        }
        status_display = {
            "pending": "⏳ 待审批", "approved": "✓ 已通过",
            "rejected": "✗ 已驳回", "skipped": "○ 待补签",
            "post_signed": "✓ 已补签",
        }

        for i, rec in enumerate(status["records"], 1):
            role_name = role_names.get(rec["role"], rec["role"])
            critical = "★ 是" if rec.get("is_critical") else "  否"
            status_text = status_display.get(rec["status"], rec["status"])
            approver = rec.get("approver", "") or "-"
            comment = rec.get("comment", "") or ""
            if len(comment) > 18:
                comment = comment[:16] + "..."

            print(f"│ {i:<6} │ {role_name:<14} │ {critical:<7} │ {status_text:<7} │ {approver:<10} {comment:<16} │")

        print("├────────┴────────────────┴─────────┴─────────┴──────────────────────────┤")

        if status.get('pending_roles'):
            pending_names = [role_names.get(r, r) for r in status['pending_roles']]
            print(f"│ 👉 当前待审批: {', '.join(pending_names)}")
        if status.get('needing_post_sign'):
            post_names = [role_names.get(r, r) for r in status['needing_post_sign']]
            print(f"│ 📝 待补签角色: {', '.join(post_names)} (使用 post-sign 命令补签)")
        if status.get('current_role') and status['release_type'] == 'regular':
            current_name = role_names.get(status['current_role'], status['current_role'])
            print(f"│ ⚠️  常规发布为串行审批，必须先由【{current_name}】审批通过")

        if status['can_enter_gray']:
            if status['is_fully_approved']:
                print("│ ✓ 全部审批通过，可以执行 release 命令发布")
            else:
                print("│ ✓ 关键角色已审批通过，可以执行 release 命令进入灰度（剩余角色需补签）")

        print("└─────────────────────────────────────────────────────────────────────────┘")

    def _print_gray_result(self, result: dict):
        print()
        print("┌─────────────────────────────────────────────────────────────────────────┐")
        print("│                       灰度发布结果                                      │")
        print("├──────┬────────────────────────────┬────────┬──────────┬─────────────────┤")
        print("│ 层级 │ 口岸                       │ 流量比 │ 状态     │ 关键时间点       │")
        print("├──────┼────────────────────────────┼────────┼──────────┼─────────────────┤")

        for r in result.get("results", []):
            tier = r.get("tier", "?")
            status = r.get("status", "?")
            status_text = "✓ 正常" if status == "passed" else "✗ 熔断回滚"
            deploy = r.get("deploy", {})
            ports = ", ".join(deploy.get("ports", []))
            ratio = f"{deploy.get('traffic_ratio', 0):.0%}"
            deploy_time = deploy.get("deployed_at", "")[11:19] if deploy.get("deployed_at") else ""
            monitor = r.get("monitor", {})

            print(f"│ {tier:<4} │ {ports[:26]:<26} │ {ratio:<6} │ {status_text:<8} │ 发布 {deploy_time}   │")

            if monitor.get("circuit_break_event"):
                cb = monitor["circuit_break_event"]
                print(f"│      │   └─ 熔断原因: {cb.get('reason', '')[:30]}")
                print(f"│      │   └─ 触发值: {cb.get('trigger_value', 0):.4f} > 阈值: {cb.get('threshold', 0):.4f}")
                print(f"│      │   └─ 熔断时间: {cb.get('triggered_at', '')[:19]}")
                rollback = r.get("rollback", {})
                if rollback:
                    print(f"│      │   └─ 回滚完成: {rollback.get('rollback_completed_at', '')[:19]}")

        print("├──────┴────────────────────────────┴────────┴──────────┴─────────────────┤")
        if result.get("circuit_break"):
            print("│ ❌ 结论: 熔断触发，已自动回滚至上一稳定版本")
            print(f"│    影响口岸: {', '.join(result.get('results', [])[-1].get('deploy', {}).get('ports', []))}")
        else:
            print("│ ✅ 结论: 全量发布成功，所有口岸已升级至新版本")
        print("└─────────────────────────────────────────────────────────────────────────┘")


# ============== CLI Commands ==============

def cmd_submit(args, platform: ReleasePlatform):
    try:
        record = platform.submit_release(
            version=args.version,
            previous_version=args.previous_version,
            release_type=args.type,
            applicant=args.applicant,
            reason=args.reason or "",
            hotfix_mode=getattr(args, "hotfix_mode", None),
        )
        print(f"\n📋 发布ID: {record.release_id}")
        print(f"   状态: {record.status.value}")
        if record.status == ReleaseStatus.CHECK_FAILED:
            print("   前置校验未通过，发布已阻断。请修复后重新提交。")
        elif record.status == ReleaseStatus.APPROVAL_PASSED:
            print("   审批已通过（事后补签模式），可执行 release 命令进行灰度发布。")
        elif record.status == ReleaseStatus.PENDING_APPROVAL:
            status = platform.get_approval_status(record.release_id)
            pending = status.get("pending_roles", [])
            if pending:
                print(f"   等待审批: {', '.join(pending)}")
    except ApprovalError as e:
        print(f"\n❌ 提交失败: {e}")
        sys.exit(1)


def cmd_approve(args, platform: ReleasePlatform):
    try:
        record = platform.approve_release(
            release_id=args.release_id,
            role=args.role,
            approver=args.approver,
            comment=args.comment or "",
        )
        print(f"\n✅ 审批完成")
        print(f"   发布ID: {record.release_id}")
        print(f"   状态: {record.status.value}")
        if record.status == ReleaseStatus.APPROVAL_PASSED:
            print("   可执行 release 命令进行灰度发布。")
    except ApprovalError:
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        sys.exit(1)


def cmd_post_sign(args, platform: ReleasePlatform):
    try:
        record = platform.post_sign(
            release_id=args.release_id,
            role=args.role,
            approver=args.approver,
            comment=args.comment or "",
        )
        print(f"\n✅ 补签完成")
        print(f"   发布ID: {record.release_id}")
        status = platform.get_approval_status(record.release_id)
        remaining = status.get("needing_post_sign", [])
        if remaining:
            print(f"   剩余待补签: {', '.join(remaining)}")
        else:
            print("   全部补签已完成 ✓")
    except ApprovalError:
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        sys.exit(1)


def cmd_reject(args, platform: ReleasePlatform):
    try:
        record = platform.reject_release(
            release_id=args.release_id,
            role=args.role,
            approver=args.approver,
            comment=args.comment or "",
        )
        print(f"\n❌ 审批已驳回")
        print(f"   发布ID: {record.release_id}")
        print(f"   状态: {record.status.value}")
    except ApprovalError:
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        sys.exit(1)


def cmd_release(args, platform: ReleasePlatform):
    try:
        result = platform.execute_gray_release(
            release_id=args.release_id,
            simulate_failure_tier=args.simulate_failure or 0,
        )
        success = result.get("success")
        print(f"\n🎯 灰度发布结果: {'✅ 成功' if success else '❌ 熔断回滚'}")
    except Exception as e:
        print(f"\n❌ 发布失败: {e}")
        sys.exit(1)


def cmd_drill(args, platform: ReleasePlatform):
    result = platform.run_drill(
        target_version=args.target_version,
        rollback_version=args.rollback_version,
    )
    status_icon = "✅" if result["status"] == "success" else "❌"
    print(f"\n{status_icon} 演练{result['status']}")
    print(f"   演练ID: {result['drill_id']}")
    print(f"   状态: {result['status']}")
    print(f"   耗时: {result['duration_seconds']}s")
    print(f"   详情: {result['detail']}")


def cmd_report(args, platform: ReleasePlatform):
    report = platform.generate_report(week_offset=args.week_offset or 0)

    print()
    print("┌─────────────────────────────────────────────────┐")
    print("│              📊 运营周报已生成                  │")
    print("├─────────────────────────────────────────────────┤")
    print(f"│ 报告ID:   {report['report_id']}")
    print(f"│ 周期:     {report['period']}")
    print(f"│ 总发布:   {report['total_releases']} 次")
    print(f"│ 成功:     {report['success_releases']} 次 ({report['success_rate']:.2%})")
    print(f"│ 回滚:     {report['rollback_count']} 次")
    print(f"│ 平均审批: {report['avg_approval_duration_minutes']} 分钟")
    print("├─────────────────────────────────────────────────┤")
    print("│              📁 生成的文件                      │")
    print("├─────────────────────────────────────────────────┤")

    file_paths = report.get("file_paths", {})
    if not file_paths:
        print("│  (未生成报表文件，可能缺少依赖)                 │")
    else:
        for fmt, path in sorted(file_paths.items()):
            full_path = os.path.abspath(path)
            icon = {
                "json": "📄", "csv": "📊", "excel": "📗", "pdf": "📕",
            }.get(fmt, "📄")
            print(f"│ {icon} {fmt.upper():<5}: {full_path}")
        charts_dir = os.path.join(os.path.dirname(file_paths.get("json", ".")), "charts")
        if os.path.isdir(charts_dir):
            report_base = report.get("report_id", "")
            for cf in sorted(os.listdir(charts_dir)):
                if report_base in cf and cf.endswith(".png"):
                    chart_full = os.path.abspath(os.path.join(charts_dir, cf))
                    print(f"│ 📈 趋势图: {chart_full}")
                    break

    if report.get('failure_rate_by_port'):
        print("├─────────────────────────────────────────────────┤")
        print("│              📍 各口岸失败率                    │")
        print("├─────────────────────────────────────────────────┤")
        for port, rate in report['failure_rate_by_port'].items():
            bar = "█" * int(rate * 50)
            print(f"│ {port:<12} {rate:>6.2%} {bar}")

    print("└─────────────────────────────────────────────────┘")
    print()


def cmd_audit(args, platform: ReleasePlatform):
    records = platform.query_audit(
        start_date=args.start_date or "",
        end_date=args.end_date or "",
        action=args.action or "",
        operator=args.operator or "",
        target=args.target or "",
    )
    print(f"\n📝 审计记录数: {len(records)}")
    for r in records[:30]:
        tamper = ""
        if r.get("tamper_check") == "TAMPERED":
            tamper = " ⚠️ TAMPERED"
        print(f"  [{r['timestamp']}] {r['action']:20s} by {r['operator']:12s} -> {r['target']:20s} {r.get('detail', '')[:40]}{tamper}")
    if len(records) > 30:
        print(f"  ... 共 {len(records)} 条，仅显示前30条")


def cmd_release_history(args, platform: ReleasePlatform):
    records = platform.query_release_history(
        version=args.version or "",
        port=args.port or "",
        start_date=args.start_date or "",
        end_date=args.end_date or "",
        release_id=args.release_id or "",
    )

    if not records:
        print("\n❌ 未找到匹配的发布记录")
        filters = []
        if args.version:
            filters.append(f"版本={args.version}")
        if args.port:
            filters.append(f"口岸={args.port}")
        if args.start_date or args.end_date:
            filters.append(f"时间={args.start_date or '*'}~{args.end_date or '*'}")
        if args.release_id:
            filters.append(f"ID={args.release_id}")
        if filters:
            filter_str = ", ".join(filters)
            print(f"   筛选条件: {filter_str}")
            if args.port:
                print(f"   💡 提示: 指定口岸「{args.port}」无匹配记录，可能该口岸尚未进行灰度部署")
        return

    print(f"\n🔍 找到 {len(records)} 条发布记录")
    print()

    reason_map = {
        "declaration_failure_rate": "报关失败率过高",
        "clearance_delay_rate": "清关延迟率过高",
        "manifest_anomaly_rate": "舱单异常率过高",
    }

    for i, rec in enumerate(records, 1):
        print(f"{'─'*70}")
        print(f"【记录 {i}】")
        print(f"  发布ID:     {rec['release_id']}")
        print(f"  版本:       {rec['version']} → {rec['previous_version']}")
        print(f"  类型:       {rec['release_type']}  {'(紧急)' if rec['release_type'] == 'hotfix' else ''}")
        print(f"  状态:       {rec['status']}")
        print(f"  申请人:     {rec['applicant']}")
        print(f"  申请时间:   {rec['apply_time']}")
        if rec.get('finish_time'):
            print(f"  完成时间:   {rec['finish_time']}")
        if rec.get('emergency_reason'):
            print(f"  紧急原因:   {rec['emergency_reason']}")

        approval = rec.get('approval', {})
        if approval.get('records'):
            status_label = {
                "pending": "待审批",
                "approved": "✅ 已通过",
                "rejected": "❌ 已驳回",
                "skipped": "⏳ 待补签",
                "post_signed": "📝 已补签",
            }
            critical_label = {True: "★关键", False: "  "}
            print(f"  审批记录:")
            for ar in approval['records']:
                s = status_label.get(ar['status'], ar['status'])
                c = critical_label.get(ar.get('is_critical'), "  ")
                approver_str = ar.get('approver', '')
                comment_str = ar.get('comment', '')[:40]
                print(f"    {c} {ar['role']:12s} {s:12s} {approver_str:16s} {comment_str}")

        ports = rec.get('port_records', [])
        if ports:
            print(f"  口岸灰度记录:")
            port_status_label = {
                "deployed": "已部署",
                "monitoring": "监控中",
                "released": "全量发布",
                "circuit_break": "已熔断",
                "rolled_back": "已回滚",
            }
            for pr in ports:
                ps = port_status_label.get(pr.get('status', ''), pr.get('status', ''))
                print(f"    ┌ {pr['port_name']}  tier={pr['tier']}  [{ps}]")
                if pr.get('deploy_started_at'):
                    print(f"    │ 灰度发布时间: {pr['deploy_started_at']}")
                if pr.get('deploy_completed_at'):
                    print(f"    │ 部署完成时间: {pr['deploy_completed_at']}")
                if pr.get('monitoring_started_at'):
                    print(f"    │ 监控开始时间: {pr['monitoring_started_at']}")
                cb = pr.get('circuit_break')
                if cb:
                    reason = reason_map.get(cb.get('reason', ''), cb.get('reason', ''))
                    print(f"    │ 熔断原因:     {reason}")
                    print(f"    │ 熔断触发值:   {cb.get('trigger_value', 0):.4f} > 阈值 {cb.get('threshold', 0):.4f}")
                    if cb.get('triggered_at'):
                        print(f"    │ 熔断时间:     {cb['triggered_at']}")
                    if cb.get('rollback_version'):
                        print(f"    │ 回滚至版本:   {cb['rollback_version']}")
                    if cb.get('rollback_completed_at'):
                        print(f"    │ 回滚完成时间: {cb['rollback_completed_at']}")
                if pr.get('rollback_completed_at') and not cb:
                    print(f"    │ 回滚完成时间: {pr['rollback_completed_at']}")
                print(f"    └")

        cbs = rec.get('circuit_break_events', [])
        if cbs:
            print(f"  熔断事件:")
            for cb in cbs:
                reason = reason_map.get(cb['reason'], cb['reason'])
                print(f"    - {cb['port_name']}: {reason} at {cb['triggered_at']}")
                print(f"      触发值 {cb['trigger_value']:.4f} > 阈值 {cb['threshold']:.4f}")
                print(f"      回滚至 {cb['rollback_version']} at {cb['rollback_completed_at']}")

    if args.export:
        output_path = os.path.abspath(args.export)
        exported = platform.export_release_history(
            output_path=output_path,
            version=args.version or "",
            port=args.port or "",
            start_date=args.start_date or "",
            end_date=args.end_date or "",
            release_id=args.release_id or "",
        )
        print(f"\n✅ 已导出到: {os.path.abspath(exported)}")

    print()


def cmd_release_summary(args, platform: ReleasePlatform):
    summary = platform.get_release_summary(
        version=args.version or "",
        start_date=args.start_date or "",
        end_date=args.end_date or "",
    )

    print()
    print("╔" + "═" * 68 + "╗")
    print("║" + " " * 27 + "📊 发布健康度汇总" + " " * 27 + "║")
    print("╠" + "═" * 68 + "╣")
    print(f"║  整体健康度: {summary['health_label']:<20}  评分: {summary['health_score']:>3}/100" + " " * 12 + "║")
    print("╠" + "═" * 68 + "╣")
    print(f"║  总发布数:     {summary['total']:<8}  常规发布: {summary['regular_count']:<5}  热修复: {summary['hotfix_count']:<5}     ║")
    print(f"║  发布成功数:   {summary['success_count']:<8}  成功率:   {summary['success_rate']:>5.1f}%" + " " * 25 + "║")
    print(f"║  熔断次数:     {summary['circuit_break_count']:<8}  回滚次数: {summary['rollback_count']:<5}" + " " * 24 + "║")
    print(f"║  补签完成率:   {summary['post_sign_rate']:>5.1f}%  ({summary['post_sign_done']}/{summary['post_sign_total']})" + " " * 30 + "║")
    print(f"║  平均审批时长: {summary['avg_approval_minutes']:.1f} 分钟" + " " * 42 + "║")
    print("╠" + "═" * 68 + "╣")

    if summary.get("by_version"):
        print("║  按版本统计:" + " " * 57 + "║")
        print("║  " + "版本        总数 成功 回滚 常规 热修" + " " * 30 + "║")
        for v, s in sorted(summary["by_version"].items()):
            print(f"║  {v:<12s} {s['count']:>3}  {s['success']:>3}  {s['rollback']:>3}  {s['regular']:>3}  {s['hotfix']:>3}" + " " * 34 + "║")

    print("╚" + "═" * 68 + "╝")

    if args.export:
        export_path = os.path.abspath(args.export)
        ext = os.path.splitext(export_path)[1].lower()

        if ext == ".json":
            with open(export_path, "w", encoding="utf-8") as f:
                json.dump(summary, f, ensure_ascii=False, indent=2)
            print(f"\n✅ 已导出JSON: {export_path}")
        elif ext in (".xlsx", ".xls"):
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment

                wb = Workbook()
                ws = wb.active
                ws.title = "健康度汇总"

                headers = ["指标", "数值"]
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="4472C4")
                    cell.alignment = Alignment(horizontal="center")

                overview_items = [
                    ("总发布数", summary["total"]),
                    ("常规发布数", summary["regular_count"]),
                    ("热修复发布数", summary["hotfix_count"]),
                    ("发布成功数", summary["success_count"]),
                    ("发布成功率(%)", summary["success_rate"]),
                    ("熔断次数", summary["circuit_break_count"]),
                    ("回滚次数", summary["rollback_count"]),
                    ("补签完成率(%)", summary["post_sign_rate"]),
                    ("平均审批时长(分钟)", summary["avg_approval_minutes"]),
                    ("健康度评分", summary["health_score"]),
                    ("健康状态", summary["health_label"]),
                ]
                for i, (k, v) in enumerate(overview_items, 2):
                    ws.cell(row=i, column=1, value=k)
                    ws.cell(row=i, column=2, value=v)

                ws2 = wb.create_sheet("按版本明细")
                headers2 = ["版本", "总数", "成功", "回滚", "常规", "热修复"]
                for col, h in enumerate(headers2, 1):
                    cell = ws2.cell(row=1, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="4472C4")

                for i, (v, s) in enumerate(sorted(summary["by_version"].items()), 2):
                    ws2.cell(row=i, column=1, value=v)
                    ws2.cell(row=i, column=2, value=s["count"])
                    ws2.cell(row=i, column=3, value=s["success"])
                    ws2.cell(row=i, column=4, value=s["rollback"])
                    ws2.cell(row=i, column=5, value=s["regular"])
                    ws2.cell(row=i, column=6, value=s["hotfix"])

                for col in range(1, 7):
                    ws2.column_dimensions[chr(64 + col)].width = 16

                wb.save(export_path)
                print(f"\n✅ 已导出Excel: {export_path}")
            except ImportError:
                print("\n⚠️  缺少 openpyxl 依赖，请先安装: pip install openpyxl")
                jpath = export_path.rsplit(".", 1)[0] + ".json"
                with open(jpath, "w", encoding="utf-8") as f:
                    json.dump(summary, f, ensure_ascii=False, indent=2)
                print(f"   已 fallback 导出JSON: {jpath}")
        else:
            print(f"\n⚠️  不支持的导出格式: {ext}，支持 .json 和 .xlsx")

    print()


def cmd_port_history(args, platform: ReleasePlatform):
    result = platform.get_port_history(args.port, limit=args.limit or 10)

    if not result["records"]:
        print(f"\n❌ 口岸「{args.port}」无灰度发布记录")
        print()
        return

    print()
    print("╔" + "═" * 68 + "╗")
    print(f"║  🏛  口岸: {result['port_name']:<30}  {result['risk_level']}" + " " * 12 + "║")
    print("╠" + "═" * 68 + "╣")
    print(f"║  当前线上版本: {str(result['current_online_version'] or 'N/A'):<30}        ║")
    print(f"║  最近部署时间: {str(result['current_deploy_time'] or 'N/A'):<30}        ║")
    print(f"║  历史发布记录: {result['total_records']} 条" + " " * 45 + "║")
    print("╠" + "═" * 68 + "╣")

    status_label = {
        "deployed": "已部署",
        "monitoring": "监控中",
        "released": "✅ 全量",
        "circuit_break": "🔥 熔断",
        "rolled_back": "🔄 已回滚",
    }
    type_label = {"regular": "常规", "hotfix": "🔥热修"}

    for pr in result["records"]:
        st = status_label.get(pr["status"], pr["status"])
        rt = type_label.get(pr["release_type"], pr["release_type"])
        cb_tag = " [熔断]" if pr["circuit_break"] else ""
        rb_tag = f" 回滚:{pr['rollback_time'][:16]}" if pr["rollback_time"] else ""
        print(f"║  {pr['release_time'][:16]}  {pr['version']:<10} t{pr['tier']} {rt} {st:<8}{cb_tag}║")
        if pr["circuit_break_reason"]:
            print(f"║    └ 原因: {pr['circuit_break_reason'][:42]}" + " " * 15 + "║")
        if rb_tag:
            print(f"║    └ {rb_tag:<58} ║")

    print("╚" + "═" * 68 + "╝")
    print()


def cmd_compliance_check(args, platform: ReleasePlatform):
    result = platform.check_compliance()

    print()
    print("╔" + "═" * 68 + "╗")
    print("║" + " " * 26 + "🛡 合规校验报告" + " " * 28 + "║")
    print("╠" + "═" * 68 + "╣")
    print(f"║  检查发布总数: {result['total_checked']:<10}  合规率: {result['compliance_rate']:>5.1f}%" + " " * 25 + "║")
    print(f"║  严重问题(high):  {result['high_count']:<30}" + " " * 24 + "║")
    print(f"║  中等问题(medium): {result['medium_count']:<30}" + " " * 24 + "║")
    print(f"║  低危问题(low):   {result['low_count']:<30}" + " " * 24 + "║")
    print("╠" + "═" * 68 + "╣")

    if not result["issues"]:
        print("║" + " " * 18 + "🎉 所有发布均符合合规要求！" + " " * 22 + "║")
    else:
        print("║  问题清单 (按严重程度排序):" + " " * 40 + "║")
        sev_label = {"high": "🔥HIGH", "medium": "⚠️ MED", "low": "  LOW"}
        for i, issue in enumerate(result["issues"], 1):
            sev = sev_label.get(issue["severity"], issue["severity"])
            print(f"║  {i:>2}. [{sev}] {issue['issue_type']}" + " " * 43 + "║")
            print(f"║      发布ID: {issue['release_id']}  版本: {issue['version']:<15}║")
            print(f"║      详情: {issue['detail'][:55]}" + " " * 2 + "║")
            print(f"║      👉 {issue['action'][:58]}" + " " * 1 + "║")

    print("╚" + "═" * 68 + "╝")
    print()


def cmd_verify(args, platform: ReleasePlatform):
    result = platform.verify_audit_integrity(date_str=args.date or "")
    print()
    print("┌─────────────────────────────────────────┐")
    print("│          🔍 审计日志完整性校验          │")
    print("├─────────────────────────────────────────┤")
    print(f"│ 检查文件数: {result['files_checked']:<29} │")
    print(f"│ 总条目数:   {result['total_entries']:<29} │")
    print(f"│ 篡改条目数: {result['tampered']:<29} │")
    print("├─────────────────────────────────────────┤")
    if result['tampered'] == 0:
        print("│ ✅ 完整性校验通过，未发现篡改记录       │")
    else:
        print("│ ❌ 存在篡改记录，请核查！               │")
    print("└─────────────────────────────────────────┘")
    print()


def cmd_full_flow(args, platform: ReleasePlatform):
    print()
    print("╔" + "═" * 68 + "╗")
    print("║" + " " * 15 + "🚀 跨境物流清关系统 - 完整发布流程演示" + " " * 16 + "║")
    print("╚" + "═" * 68 + "╝")
    print()

    # ========== 场景A: 常规发布 ==========
    print("📌 场景A: 常规发布 - 前置校验 → 自动审批 → 灰度发布")
    print("─" * 70)

    try:
        record_a = platform.submit_release(
            version=args.version,
            previous_version=args.previous_version,
            release_type="regular",
            applicant="demo_user",
            reason="季度功能迭代",
        )
    except Exception as e:
        print(f"❌ 提交失败: {e}")
        record_a = None

    if record_a and record_a.status == ReleaseStatus.CHECK_FAILED:
        print("\n" + "╔" + "═" * 68 + "╗")
        print("║" + " " * 20 + "❌ 前置校验未通过，演示终止" + " " * 20 + "║")
        print("╚" + "═" * 68 + "╝")
        print()
        print("💡 请根据上方阻断指标修复问题后，重新执行演示。")
        print()
        return
    elif record_a and record_a.status in (ReleaseStatus.CHECK_PASSED, ReleaseStatus.PENDING_APPROVAL):
        print("\n✅ 前置校验通过，开始自动审批...")
        try:
            result_a = platform.auto_approve_and_release(record_a.release_id)
            if result_a.get("success"):
                release_ok = result_a["release_result"].get("success")
                print(f"\n✅ 场景A完成: {'全量发布成功 ✓' if release_ok else '熔断触发，已自动回滚'}")
            else:
                print(f"\n❌ 场景A未能完成: {result_a.get('reason')}")
                print("\n⚠️  常规发布流程失败，跳过后续场景。")
                return
        except Exception as e:
            print(f"\n❌ 场景A执行失败: {e}")
            print("\n⚠️  常规发布流程异常，跳过后续场景。")
            return
    elif record_a:
        print(f"\n⚠️  场景A状态异常: {record_a.status.value}")

    # ========== 场景B: 紧急热修复 ==========
    print()
    print("📌 场景B: 紧急热修复 - 并行审批 + 模拟熔断回滚 + 事后补签")
    print("─" * 70)

    try:
        record_b = platform.submit_release(
            version=f"{args.version}-hotfix",
            previous_version=args.previous_version,
            release_type="hotfix",
            applicant="admin",
            reason="紧急修复深圳口岸报关接口异常",
            hotfix_mode="parallel",
        )
    except Exception as e:
        print(f"❌ 提交失败: {e}")
        record_b = None

    if record_b and record_b.status in (ReleaseStatus.CHECK_PASSED, ReleaseStatus.PENDING_APPROVAL):
        print("\n关键角色(关务/技术)并行审批...")
        try:
            for role in ["customs", "tech"]:
                record_b = platform.approve_release(
                    release_id=record_b.release_id,
                    role=role,
                    approver=f"{role}_reviewer",
                    comment=f"[紧急并行审批] {role}评估通过",
                )
        except Exception as e:
            print(f"❌ 审批失败: {e}")

        if record_b.status == ReleaseStatus.APPROVAL_PASSED:
            print("\n✅ 关键角色审批通过，直接进入灰度发布（非关键角色后续补签）...")
            try:
                result_b = platform.execute_gray_release(
                    record_b.release_id,
                    simulate_failure_tier=2,
                )
                cb = result_b.get("circuit_break")
                print(f"\n🔥 场景B完成: {'熔断触发，已自动回滚' if cb else '全量发布成功'}")

                if cb:
                    print("\n📝 对剩余角色进行事后补签...")
                    for role in ["operations", "finance"]:
                        try:
                            platform.post_sign(
                                release_id=record_b.release_id,
                                role=role,
                                approver=f"{role}_reviewer",
                                comment=f"[事后补签] 确认熔断回滚过程合规",
                            )
                        except Exception as e:
                            print(f"  ⚠️  {role}补签: {e}")
                    print("✅ 事后补签完成")
            except Exception as e:
                print(f"\n❌ 灰度发布失败: {e}")

    # ========== 场景C: 回滚演练 ==========
    print()
    print("📌 场景C: 定期回滚演练")
    print("─" * 70)
    drill_result = platform.run_drill(
        target_version=args.version,
        rollback_version=args.previous_version,
    )
    icon = "✅" if drill_result["status"] == "success" else "❌"
    print(f"\n{icon} 场景C完成: 演练{drill_result['status']}, 耗时{drill_result['duration_seconds']}s")

    # ========== 场景D: 生成周报 ==========
    print()
    print("📌 场景D: 生成运营周报 (PDF/Excel/CSV)")
    print("─" * 70)
    report = platform.generate_report()

    print()
    print("┌─────────────────────────────────────────┐")
    print("│    📊 周报已生成，文件列表:             │")
    print("├─────────────────────────────────────────┤")
    for fmt, path in sorted(report.get("file_paths", {}).items()):
        full_path = os.path.abspath(path)
        icon = {"json": "📄", "csv": "📊", "excel": "📗", "pdf": "📕"}.get(fmt, "📄")
        print(f"│ {icon} {fmt.upper():<5}: {full_path}")
    report_id = report.get("report_id", "")
    charts_dir = os.path.join("reports", "weekly", "charts")
    if os.path.isdir(charts_dir):
        for cf in sorted(os.listdir(charts_dir)):
            if report_id in cf and cf.endswith(".png"):
                chart_full = os.path.abspath(os.path.join(charts_dir, cf))
                print(f"│ 📈 趋势图: {chart_full}")
                break
    print("└─────────────────────────────────────────┘")

    print()
    print("╔" + "═" * 68 + "╗")
    print("║" + " " * 27 + "🎉 全流程演示完成" + " " * 28 + "║")
    print("╚" + "═" * 68 + "╝")
    print()


# ============== Main ==============

def main():
    parser = argparse.ArgumentParser(
        description="跨境物流清关系统版本发布与合规回滚自动化平台",
    )
    parser.add_argument("--config", default="config.yaml", help="配置文件路径")
    subparsers = parser.add_subparsers(dest="command", help="可用命令")

    p_submit = subparsers.add_parser("submit", help="提交发布申请")
    p_submit.add_argument("--version", required=True, help="发布版本号")
    p_submit.add_argument("--previous-version", required=True, help="上一稳定版本号")
    p_submit.add_argument("--type", choices=["regular", "hotfix"], default="regular", help="发布类型")
    p_submit.add_argument("--applicant", default="developer", help="申请人")
    p_submit.add_argument("--reason", default="", help="发布原因(热修复必填)")
    p_submit.add_argument("--hotfix-mode", choices=["parallel", "post_sign"], help="热修复模式(仅hotfix)")

    p_approve = subparsers.add_parser("approve", help="审批通过 (严格串行)")
    p_approve.add_argument("--release-id", required=True, help="发布ID")
    p_approve.add_argument("--role", required=True, choices=["customs", "operations", "finance", "tech"], help="审批角色")
    p_approve.add_argument("--approver", required=True, help="审批人")
    p_approve.add_argument("--comment", default="", help="审批意见")

    p_post_sign = subparsers.add_parser("post-sign", help="事后补签 (仅hotfix)")
    p_post_sign.add_argument("--release-id", required=True, help="发布ID")
    p_post_sign.add_argument("--role", required=True, choices=["customs", "operations", "finance", "tech"], help="补签角色")
    p_post_sign.add_argument("--approver", required=True, help="补签人")
    p_post_sign.add_argument("--comment", default="", help="补签意见")

    p_reject = subparsers.add_parser("reject", help="审批驳回")
    p_reject.add_argument("--release-id", required=True, help="发布ID")
    p_reject.add_argument("--role", required=True, choices=["customs", "operations", "finance", "tech"], help="审批角色")
    p_reject.add_argument("--approver", required=True, help="审批人")
    p_reject.add_argument("--comment", default="", help="驳回原因")

    p_release = subparsers.add_parser("release", help="执行灰度发布")
    p_release.add_argument("--release-id", required=True, help="发布ID")
    p_release.add_argument("--simulate-failure", type=int, default=0, help="模拟第N层熔断(测试用)")

    p_drill = subparsers.add_parser("drill", help="执行回滚演练")
    p_drill.add_argument("--target-version", default="v2.0.0", help="演练目标版本")
    p_drill.add_argument("--rollback-version", default="v1.9.0", help="演练回滚版本")

    p_report = subparsers.add_parser("report", help="生成运营周报 (PDF/Excel/CSV)")
    p_report.add_argument("--week-offset", type=int, default=0, help="周偏移(0=本周)")

    p_audit = subparsers.add_parser("audit", help="查询审计日志")
    p_audit.add_argument("--start-date", default="", help="开始日期 YYYY-MM-DD")
    p_audit.add_argument("--end-date", default="", help="结束日期 YYYY-MM-DD")
    p_audit.add_argument("--action", default="", help="操作类型")
    p_audit.add_argument("--operator", default="", help="操作人")
    p_audit.add_argument("--target", default="", help="目标对象")

    p_history = subparsers.add_parser("release-history", help="按版本/口岸/时间查询发布回滚记录")
    p_history.add_argument("--version", default="", help="版本号")
    p_history.add_argument("--port", default="", help="口岸名称")
    p_history.add_argument("--release-id", default="", help="发布ID")
    p_history.add_argument("--start-date", default="", help="开始日期 YYYY-MM-DD")
    p_history.add_argument("--end-date", default="", help="结束日期 YYYY-MM-DD")
    p_history.add_argument("--export", default="", help="导出路径 (.json 或 .csv)")

    p_summary = subparsers.add_parser("release-summary", help="发布健康度汇总 (支持JSON/Excel导出)")
    p_summary.add_argument("--version", default="", help="按版本号过滤")
    p_summary.add_argument("--start-date", default="", help="开始日期 YYYY-MM-DD")
    p_summary.add_argument("--end-date", default="", help="结束日期 YYYY-MM-DD")
    p_summary.add_argument("--export", default="", help="导出路径 (.json 或 .xlsx)")

    p_port = subparsers.add_parser("port-history", help="按口岸聚合查询灰度发布历史")
    p_port.add_argument("--port", required=True, help="口岸名称 (如: 深圳口岸)")
    p_port.add_argument("--limit", type=int, default=10, help="显示最近N条记录")

    p_compliance = subparsers.add_parser("compliance-check", help="合规校验: 审批闭环/补签超期/顺序异常/回滚未复盘")

    p_verify = subparsers.add_parser("verify", help="校验审计日志完整性")
    p_verify.add_argument("--date", default="", help="指定日期(空=全部)")

    p_full = subparsers.add_parser("full-flow", help="完整流程演示")
    p_full.add_argument("--version", default="v2.1.0", help="发布版本号")
    p_full.add_argument("--previous-version", default="v2.0.0", help="上一稳定版本号")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return

    platform = ReleasePlatform(args.config)

    commands_map = {
        "submit": cmd_submit,
        "approve": cmd_approve,
        "post-sign": cmd_post_sign,
        "reject": cmd_reject,
        "release": cmd_release,
        "drill": cmd_drill,
        "report": cmd_report,
        "audit": cmd_audit,
        "release-history": cmd_release_history,
        "release-summary": cmd_release_summary,
        "port-history": cmd_port_history,
        "compliance-check": cmd_compliance_check,
        "verify": cmd_verify,
        "full-flow": cmd_full_flow,
    }

    handler = commands_map.get(args.command)
    if handler:
        handler(args, platform)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
