from __future__ import annotations

import csv
import json
import os
import logging
from datetime import datetime, timedelta
from typing import Optional
from collections import defaultdict

from models import (
    ReleaseStatus,
    WeeklyReport,
    generate_id,
    now_iso,
)

logger = logging.getLogger(__name__)


class ReportEngine:
    def __init__(self, config: dict):
        self.config = config.get("report", {})
        self.system_config = config.get("system", {})
        self.data_dir = self.system_config.get("data_dir", "./data")
        self.report_dir = self.config.get("output_dir", "./reports/weekly")
        self.trend_weeks = self.config.get("trend_weeks", 8)
        os.makedirs(self.report_dir, exist_ok=True)
        os.makedirs(os.path.join(self.report_dir, "charts"), exist_ok=True)

    def generate_weekly_report(self, week_offset: int = 0) -> WeeklyReport:
        now = datetime.now()
        monday = now - timedelta(days=now.weekday()) + timedelta(weeks=week_offset)
        sunday = monday + timedelta(days=6)
        period_start = monday.strftime("%Y-%m-%d")
        period_end = sunday.strftime("%Y-%m-%d")

        report = WeeklyReport(
            report_id=generate_id("rpt_"),
            period_start=period_start,
            period_end=period_end,
        )

        release_records = self._load_release_records(period_start, period_end)
        report.total_releases = len(release_records)
        report.success_releases = len([
            r for r in release_records
            if r.get("status") in (ReleaseStatus.FULL_RELEASED.value, ReleaseStatus.GRAY_MONITORING.value)
        ])
        report.rollback_count = len([
            r for r in release_records
            if r.get("status") in (ReleaseStatus.ROLLED_BACK.value, ReleaseStatus.CIRCUIT_BREAK.value)
        ])

        durations = []
        for r in release_records:
            if "approval_duration_minutes" in r:
                durations.append(r["approval_duration_minutes"])
        report.approval_duration_list = durations
        report.avg_approval_duration_minutes = round(sum(durations) / len(durations), 2) if durations else 0.0

        port_failures = defaultdict(int)
        port_totals = defaultdict(int)
        for r in release_records:
            for port, failed in r.get("port_results", {}).items():
                port_totals[port] += 1
                if failed:
                    port_failures[port] += 1
        report.failure_rate_by_port = {
            port: round(port_failures[port] / port_totals[port], 4)
            for port in port_totals
        }

        report.trend_data = self._collect_trend_data(week_offset)

        self._save_report(report)
        report.file_paths["json"] = os.path.join(self.report_dir, f"{report.report_id}.json")
        self._export_csv(report)
        report.file_paths["csv"] = os.path.join(self.report_dir, f"{report.report_id}.csv")

        excel_path = self._export_excel(report)
        if excel_path:
            report.file_paths["excel"] = excel_path

        pdf_path = self._export_pdf(report)
        if pdf_path:
            report.file_paths["pdf"] = pdf_path

        self._save_report(report)
        logger.info("周报已生成: %s (%s ~ %s)", report.report_id, period_start, period_end)
        for fmt, path in report.file_paths.items():
            logger.info("  %s: %s", fmt.upper(), path)

        return report

    def _collect_trend_data(self, current_week_offset: int) -> list[dict]:
        trend_data = []
        for w in range(self.trend_weeks - 1, -1, -1):
            offset = current_week_offset - w
            now = datetime.now()
            monday = now - timedelta(days=now.weekday()) + timedelta(weeks=offset)
            sunday = monday + timedelta(days=6)
            period_start = monday.strftime("%Y-%m-%d")
            period_end = sunday.strftime("%Y-%m-%d")

            records = self._load_release_records(period_start, period_end)
            total = len(records)
            success = len([
                r for r in records
                if r.get("status") in (ReleaseStatus.FULL_RELEASED.value, ReleaseStatus.GRAY_MONITORING.value)
            ])
            rollback = len([
                r for r in records
                if r.get("status") in (ReleaseStatus.ROLLED_BACK.value, ReleaseStatus.CIRCUIT_BREAK.value)
            ])
            durations = [r.get("approval_duration_minutes", 0) for r in records if "approval_duration_minutes" in r]
            avg_duration = round(sum(durations) / len(durations), 2) if durations else 0.0
            success_rate = round(success / total, 4) if total > 0 else 0.0

            trend_data.append({
                "week": monday.strftime("%m-%d"),
                "period": f"{period_start} ~ {period_end}",
                "total": total,
                "success": success,
                "rollback": rollback,
                "success_rate": success_rate,
                "avg_approval_duration": avg_duration,
            })
        return trend_data

    def _load_release_records(self, start: str, end: str) -> list[dict]:
        releases_dir = os.path.join(self.data_dir, "releases")
        if not os.path.exists(releases_dir):
            return []
        records = []
        for filename in os.listdir(releases_dir):
            if not filename.endswith(".json"):
                continue
            path = os.path.join(releases_dir, filename)
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            apply_time = data.get("apply_time", "")
            if start <= apply_time[:10] <= end:
                records.append(data)
        return records

    def _save_report(self, report: WeeklyReport):
        data = {
            "report_id": report.report_id,
            "period_start": report.period_start,
            "period_end": report.period_end,
            "total_releases": report.total_releases,
            "success_releases": report.success_releases,
            "rollback_count": report.rollback_count,
            "avg_approval_duration_minutes": report.avg_approval_duration_minutes,
            "approval_duration_list": report.approval_duration_list,
            "failure_rate_by_port": report.failure_rate_by_port,
            "trend_data": report.trend_data,
            "file_paths": report.file_paths,
            "success_rate": (
                round(report.success_releases / report.total_releases, 4)
                if report.total_releases > 0 else 0.0
            ),
            "generated_at": now_iso(),
        }
        path = os.path.join(self.report_dir, f"{report.report_id}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _export_csv(self, report: WeeklyReport):
        path = os.path.join(self.report_dir, f"{report.report_id}.csv")
        success_rate = (
            round(report.success_releases / report.total_releases, 4)
            if report.total_releases > 0 else 0.0
        )
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["指标", "值"])
            writer.writerow(["报告周期", f"{report.period_start} ~ {report.period_end}"])
            writer.writerow(["总发布次数", report.total_releases])
            writer.writerow(["成功发布次数", report.success_releases])
            writer.writerow(["回滚次数", report.rollback_count])
            writer.writerow(["发布成功率", f"{success_rate:.2%}"])
            writer.writerow(["平均审批时长(分钟)", report.avg_approval_duration_minutes])
            for port, rate in report.failure_rate_by_port.items():
                writer.writerow([f"{port} 失败率", f"{rate:.2%}"])
            writer.writerow([])
            writer.writerow(["趋势数据"])
            writer.writerow(["周", "总发布", "成功", "回滚", "成功率", "平均审批时长"])
            for td in report.trend_data:
                writer.writerow([
                    td["week"],
                    td["total"],
                    td["success"],
                    td["rollback"],
                    f"{td['success_rate']:.2%}",
                    td["avg_approval_duration"],
                ])
        logger.info("CSV报告已导出: %s", path)
        return path

    def _export_excel(self, report: WeeklyReport) -> Optional[str]:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import LineChart, Reference, BarChart
            from openpyxl.chart.label import DataLabelList
        except ImportError:
            logger.warning("openpyxl 未安装，跳过 Excel 导出")
            return None

        path = os.path.join(self.report_dir, f"{report.report_id}.xlsx")
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "周报概览"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        ws1["A1"] = "跨境物流清关系统 - 版本发布周报"
        ws1["A1"].font = Font(bold=True, size=16)
        ws1.merge_cells("A1:B1")

        success_rate = (
            round(report.success_releases / report.total_releases, 4)
            if report.total_releases > 0 else 0.0
        )

        data = [
            ["报告周期", f"{report.period_start} ~ {report.period_end}"],
            ["总发布次数", report.total_releases],
            ["成功发布次数", report.success_releases],
            ["回滚次数", report.rollback_count],
            ["发布成功率", f"{success_rate:.2%}"],
            ["平均审批时长(分钟)", report.avg_approval_duration_minutes],
        ]

        for i, (key, val) in enumerate(data, start=3):
            ws1.cell(row=i, column=1, value=key).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=val)
            for col in [1, 2]:
                ws1.cell(row=i, column=col).border = thin_border

        if report.failure_rate_by_port:
            ws1.cell(row=10, column=1, value="各口岸失败率").font = Font(bold=True)
            ws1.cell(row=10, column=2, value="失败率").font = Font(bold=True)
            for i, (port, rate) in enumerate(report.failure_rate_by_port.items(), start=11):
                ws1.cell(row=i, column=1, value=port)
                ws1.cell(row=i, column=2, value=f"{rate:.2%}")

        ws2 = wb.create_sheet("趋势数据")
        ws2["A1"] = "周"
        ws2["B1"] = "总发布"
        ws2["C1"] = "成功"
        ws2["D1"] = "回滚"
        ws2["E1"] = "成功率"
        ws2["F1"] = "平均审批时长"
        for col in range(1, 7):
            cell = ws2.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for i, td in enumerate(report.trend_data, start=2):
            ws2.cell(row=i, column=1, value=td["week"]).border = thin_border
            ws2.cell(row=i, column=2, value=td["total"]).border = thin_border
            ws2.cell(row=i, column=3, value=td["success"]).border = thin_border
            ws2.cell(row=i, column=4, value=td["rollback"]).border = thin_border
            ws2.cell(row=i, column=5, value=td["success_rate"]).border = thin_border
            ws2.cell(row=i, column=6, value=td["avg_approval_duration"]).border = thin_border

        chart1 = LineChart()
        chart1.title = "发布成功率趋势"
        chart1.y_axis.title = "成功率"
        chart1.x_axis.title = "周"
        chart1.height = 10
        chart1.width = 20
        data = Reference(ws2, min_col=5, min_row=1, max_row=len(report.trend_data) + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(report.trend_data) + 1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.dataLabels = DataLabelList(showVal=True, numFmt='0.00%')
        ws2.add_chart(chart1, "H1")

        chart2 = BarChart()
        chart2.title = "发布/回滚次数趋势"
        chart2.y_axis.title = "次数"
        chart2.x_axis.title = "周"
        chart2.height = 10
        chart2.width = 20
        data = Reference(ws2, min_col=2, min_row=1, max_col=4, max_row=len(report.trend_data) + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(report.trend_data) + 1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.dataLabels = DataLabelList(showVal=True)
        ws2.add_chart(chart2, "H20")

        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 40
        ws2.column_dimensions["A"].width = 12
        for col in ["B", "C", "D", "E", "F"]:
            ws2.column_dimensions[col].width = 15

        wb.save(path)
        logger.info("Excel报告已导出: %s", path)
        return path

    def _export_pdf(self, report: WeeklyReport) -> Optional[str]:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            )
            from reportlab.graphics.shapes import Drawing, Line, Rect, String
            from reportlab.graphics.charts.barcharts import VerticalBarChart
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
        except ImportError:
            logger.warning("reportlab 未安装，跳过 PDF 导出")
            return None

        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import matplotlib.font_manager as fm
            has_matplotlib = True
        except ImportError:
            has_matplotlib = False

        path = os.path.join(self.report_dir, f"{report.report_id}.pdf")

        doc = SimpleDocTemplate(
            path,
            pagesize=A4,
            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30,
        )
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"], fontSize=18, spaceAfter=12, textColor=colors.HexColor("#4472C4")
        )
        heading_style = ParagraphStyle(
            "CustomHeading", parent=styles["Heading2"], fontSize=14, spaceAfter=8,
            textColor=colors.HexColor("#2E75B6")
        )
        normal_style = styles["Normal"]

        elements.append(Paragraph("跨境物流清关系统 - 版本发布周报", title_style))
        elements.append(Paragraph(
            f"报告周期: {report.period_start} ~ {report.period_end}",
            styles["Italic"]
        ))
        elements.append(Spacer(1, 12))

        success_rate = (
            round(report.success_releases / report.total_releases, 4)
            if report.total_releases > 0 else 0.0
        )

        summary_data = [
            ["指标", "数值"],
            ["总发布次数", str(report.total_releases)],
            ["成功发布次数", str(report.success_releases)],
            ["回滚次数", str(report.rollback_count)],
            ["发布成功率", f"{success_rate:.2%}"],
            ["平均审批时长", f"{report.avg_approval_duration_minutes} 分钟"],
        ]
        summary_table = Table(summary_data, colWidths=[2.5 * inch, 3 * inch])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F2F2F2")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
            ("PADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 16))

        if has_matplotlib:
            chart_path = self._generate_trend_chart(report)
            if chart_path:
                from reportlab.platypus import Image
                elements.append(Paragraph("核心指标趋势", heading_style))
                img = Image(chart_path, width=7 * inch, height=3.5 * inch)
                elements.append(img)
                elements.append(Spacer(1, 12))

        if report.failure_rate_by_port:
            elements.append(Paragraph("各口岸失败率", heading_style))
            port_data = [["口岸", "失败率"]]
            for port, rate in report.failure_rate_by_port.items():
                port_data.append([port, f"{rate:.2%}"])
            port_table = Table(port_data, colWidths=[2.5 * inch, 2 * inch])
            port_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ("PADDING", (0, 0), (-1, -1), 8),
            ]))
            elements.append(port_table)
            elements.append(Spacer(1, 12))

        elements.append(Paragraph("最近8周趋势明细", heading_style))
        trend_header = ["周", "总发布", "成功", "回滚", "成功率", "审批时长"]
        trend_data = [trend_header]
        for td in report.trend_data:
            trend_data.append([
                td["week"],
                str(td["total"]),
                str(td["success"]),
                str(td["rollback"]),
                f"{td['success_rate']:.2%}",
                f"{td['avg_approval_duration']} 分钟",
            ])
        trend_table = Table(trend_data, colWidths=[0.9 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch, 1 * inch, 1.2 * inch])
        trend_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("PADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
        ]))
        elements.append(trend_table)

        doc.build(elements)
        logger.info("PDF报告已导出: %s", path)
        return path

    def _generate_trend_chart(self, report: WeeklyReport) -> Optional[str]:
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
        except ImportError:
            return None

        chart_path = os.path.join(self.report_dir, "charts", f"{report.report_id}_trend.png")

        weeks = [td["week"] for td in report.trend_data]
        success_rates = [td["success_rate"] * 100 for td in report.trend_data]
        rollback_counts = [td["rollback"] for td in report.trend_data]
        approval_durations = [td["avg_approval_duration"] for td in report.trend_data]

        fig, ax1 = plt.subplots(figsize=(10, 5))

        color1 = "#2E75B6"
        color2 = "#C00000"
        color3 = "#70AD47"

        ax1.set_xlabel("周")
        ax1.set_ylabel("发布成功率 (%)", color=color1)
        line1 = ax1.plot(weeks, success_rates, color=color1, marker="o", linewidth=2, label="发布成功率")
        ax1.tick_params(axis="y", labelcolor=color1)
        ax1.set_ylim([0, 105])
        for i, v in enumerate(success_rates):
            ax1.annotate(f"{v:.1f}%", (weeks[i], v + 1), ha="center", fontsize=8, color=color1)

        ax2 = ax1.twinx()
        ax2.set_ylabel("回滚次数 / 审批时长", color=color2)
        bar1 = ax2.bar(
            [w for w in weeks], rollback_counts, alpha=0.6, color=color2,
            label="回滚次数", width=0.35, align="edge"
        )
        line2 = ax2.plot(
            weeks, approval_durations, color=color3, marker="s", linewidth=2,
            label="平均审批时长(分钟)"
        )
        ax2.tick_params(axis="y", labelcolor=color2)

        for i, v in enumerate(rollback_counts):
            if v > 0:
                ax2.annotate(str(v), (i, v + 0.2), ha="center", fontsize=8, color=color2)

        for i, v in enumerate(approval_durations):
            ax2.annotate(f"{v:.0f}", (weeks[i], v + 0.3), ha="center", fontsize=8, color=color3)

        lines = line1 + line2
        labels = [l.get_label() for l in lines]
        ax1.legend(lines, labels, loc="upper left")
        ax2.legend([bar1], ["回滚次数"], loc="upper right")

        plt.title("发布成功率、回滚次数、审批时长趋势")
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.savefig(chart_path, dpi=150, bbox_inches="tight")
        plt.close()

        logger.info("趋势图已生成: %s", chart_path)
        return chart_path

    def list_reports(self) -> list[dict]:
        reports = []
        if not os.path.exists(self.report_dir):
            return reports
        for filename in os.listdir(self.report_dir):
            if not filename.endswith(".json"):
                continue
            path = os.path.join(self.report_dir, filename)
            with open(path, "r", encoding="utf-8") as f:
                reports.append(json.load(f))
        return sorted(reports, key=lambda r: r.get("period_start", ""), reverse=True)
